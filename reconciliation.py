"""
财务多系统对账 — ERP / 银行流水 / 第三方支付 三账合一
优化版：精确匹配 + 金额差异检测 + 合并支付识别
"""

import csv, os
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DATA_DIR = r"D:\zhuomian\企业复杂难题例题\data\对账"
OUTPUT_DIR = r"D:\zhuomian\claude答案"
os.makedirs(OUTPUT_DIR, exist_ok=True)


def read_csv(filename):
    path = os.path.join(DATA_DIR, filename)
    with open(path, "r", encoding="utf-8-sig") as f:
        return list(csv.DictReader(f))


def amt(val):
    return 0.0 if val is None or val.strip() == "" else float(val.replace(",", ""))


erp = read_csv("erp_sales_202604.csv")
bank = read_csv("bank_statement_202604.csv")
third = read_csv("third_payment_202604.csv")

print(f"ERP:{len(erp)}  银行:{len(bank)}  第三方:{len(third)}")

# ── 统一格式 ──
for r in erp:
    r["_amt"] = amt(r["金额"])
    r["_date"] = r["下单日期"].strip()
for r in bank:
    r["_amt"] = amt(r.get("收入金额", "0")) - amt(r.get("支出金额", "0"))
    r["_date"] = r["交易日期"].strip()
for r in third:
    r["_amt"] = amt(r["金额"])
    r["_fee"] = amt(r["手续费"])
    r["_net"] = amt(r["实收金额"])
    r["_date"] = r["交易时间"][:10].strip() if r.get("交易时间") else ""

# ── 索引 ──
erp_by_oid = {r["订单号"]: r for r in erp}
third_by_oid = defaultdict(list)
for r in third:
    third_by_oid[r["商户订单号"].strip()].append(r)

# 按日期索引 ERP
erp_by_date = defaultdict(list)
for r in erp:
    erp_by_date[r["_date"]].append(r)

diffs = []  # 差异明细


def add_diff(typ, src, ref, ea, ba, ta, note):
    diffs.append(dict(zip(
        ["差异类型", "数据来源", "参考号", "ERP金额", "银行金额", "第三方金额", "差异说明"],
        [typ, src, ref, round(ea, 2), round(ba, 2), round(ta, 2), note])))


# ═══════════════════════════════════════════
# 1. ERP × 第三方支付 对账
# ═══════════════════════════════════════════

erp_matched_third = set()
for oid, erec in erp_by_oid.items():
    trecs = third_by_oid.get(oid, [])
    if not trecs:
        if erec["支付方式"] in ("支付宝", "微信支付"):
            add_diff("第三方缺失", "ERP", oid, erec["_amt"], 0, 0,
                     f"ERP有订单{oid}但第三方无记录（{erec['支付方式']}）")
        continue
    erp_matched_third.add(oid)
    for tr in trecs:
        if tr["状态"] == "退款":
            add_diff("退款", "第三方", oid, 0, 0, tr["_amt"], f"退款{tr['_amt']}元 {tr['交易号']}")
            continue
        diff = round(erec["_amt"] - tr["_amt"], 2)
        if abs(diff) > 0.01:
            add_diff("金额差异(ERP×第三方)", f"ERP+第三方", oid, erec["_amt"], 0, tr["_amt"],
                     f"ERP{erec['_amt']}≠第三方{tr['_amt']} 差异{diff}")

# 第三方有但 ERP 没有（通过商户订单号）
for r in third:
    oid = r["商户订单号"].strip()
    if oid not in erp_by_oid and oid != "REFUND-0420":
        add_diff("ERP缺失(第三方有)", "第三方", oid, 0, 0, r["_amt"],
                 f"第三方有订单{oid}但ERP无对应记录")
    elif oid == "REFUND-0420":
        add_diff("退款", "第三方", oid, 0, 0, r["_amt"], f"退款{r['_amt']}元 {r['交易号']}")

# 重复记录
oid_counts = defaultdict(int)
for r in third:
    oid_counts[r["商户订单号"].strip()] += 1
for oid, n in oid_counts.items():
    if n > 1 and oid != "REFUND-0420":
        add_diff("重复记录", "第三方", oid, 0, 0, 0, f"商户订单号{oid}出现{n}次")


# ═══════════════════════════════════════════
# 2. ERP × 银行流水 对账
# ═══════════════════════════════════════════

erp_matched_bank = set()
bank_matched = set()

for bi, brec in enumerate(bank):
    b_amt = brec["_amt"]
    b_date = brec["_date"]
    if b_amt <= 0:
        continue

    # 找出同一天所有未匹配的 ERP 订单
    candidates = [r for r in erp_by_date.get(b_date, [])
                  if r["订单号"] not in erp_matched_bank]

    # 精确匹配：金额完全一致
    exact = [r for r in candidates if abs(r["_amt"] - b_amt) < 0.01]

    if len(exact) == 1:
        # ✅ 完全匹配
        erp_matched_bank.add(exact[0]["订单号"])
        bank_matched.add(bi)
        continue

    elif len(exact) > 1:
        # 同金额多条 → 标记重复风险
        for r in exact:
            erp_matched_bank.add(r["订单号"])
        bank_matched.add(bi)
        add_diff("重复匹配", "银行", f"银行#{bi+1}", round(sum(r["_amt"] for r in exact),2),
                 b_amt, 0, f"银行{b_amt}匹配到{len(exact)}条ERP订单")
        continue

    # 没有精确匹配 → 检查是否合并支付（银行一笔 = 多笔 ERP 合计）
    # 尝试找到一组 ERP 订单合计等于银行金额
    if len(candidates) >= 2:
        total = sum(r["_amt"] for r in candidates)
        # 先试所有候选订单合计
        combos = []
        # 从简单开始：试 2 个候选订单组合
        from itertools import combinations
        found_combo = False
        for n in range(2, min(len(candidates) + 1, 6)):
            for combo in combinations(candidates, n):
                s = sum(r["_amt"] for r in combo)
                if abs(s - b_amt) < 0.01:
                    for r in combo:
                        erp_matched_bank.add(r["订单号"])
                    bank_matched.add(bi)
                    oids = ",".join(r["订单号"] for r in combo)
                    add_diff("合并支付", "银行", f"银行#{bi+1}", round(s,2), b_amt, 0,
                             f"银行一笔{b_amt}=ERP多笔合计:{oids}")
                    found_combo = True
                    break
            if found_combo:
                break

    if bi not in bank_matched and candidates:
        # 有候选但金额不匹配 → 金额差异
        if len(candidates) == 1:
            erp_matched_bank.add(candidates[0]["订单号"])
            bank_matched.add(bi)
            diff = round(candidates[0]["_amt"] - b_amt, 2)
            add_diff("金额差异(ERP×银行)", "ERP+银行", f"{candidates[0]['订单号']}/银行#{bi+1}",
                     candidates[0]["_amt"], b_amt, 0,
                     f"ERP{candidates[0]['_amt']}≠银行{b_amt} 差异{diff}")
        else:
            # 多候选但无法组合匹配
            bank_matched.add(bi)
            total_cand = round(sum(r["_amt"] for r in candidates), 2)
            add_diff("金额差异(ERP×银行)", "ERP+银行", f"银行#{bi+1}",
                     total_cand, b_amt, 0,
                     f"当天ERP订单合计{total_cand}≠银行{b_amt}")
    elif bi not in bank_matched:
        # 当天无 ERP 订单
        bank_matched.add(bi)
        add_diff("ERP缺失(银行有)", "银行", f"银行#{bi+1}", 0, b_amt, 0,
                 f"银行{b_date}有收款{b_amt}但ERP无对应订单")

# ERP 有但银行没有（银行转账的订单）
for oid, erec in erp_by_oid.items():
    if oid not in erp_matched_bank and erec["支付方式"] == "银行转账":
        add_diff("银行缺失(ERP有)", "ERP", oid, erec["_amt"], 0, 0,
                 f"ERP订单{oid}({erec['_amt']}元,银行转账)但银行流水无对应记录")


# ═══════════════════════════════════════════
# 3. 分类汇总 & 排序
# ═══════════════════════════════════════════

type_summary = defaultdict(lambda: {"n": 0, "ea": 0.0, "ba": 0.0, "ta": 0.0})
for d in diffs:
    t = d["差异类型"]
    type_summary[t]["n"] += 1
    type_summary[t]["ea"] += d["ERP金额"]
    type_summary[t]["ba"] += d["银行金额"]
    type_summary[t]["ta"] += d["第三方金额"]

# 超 50 条按金额降序
if len(diffs) > 50:
    diffs.sort(key=lambda d: abs(d["ERP金额"]) + abs(d["银行金额"]) + abs(d["第三方金额"]), reverse=True)


# ═══════════════════════════════════════════
# 4. 调节表：按支付方式汇总
# ═══════════════════════════════════════════

adj = defaultdict(lambda: {"erp": 0.0, "bank": 0.0, "third": 0.0, "fee": 0.0})
for r in erp:
    adj[r["支付方式"]]["erp"] += r["_amt"]
# 银行流水按支付方式归类
for bi, brec in enumerate(bank):
    if bi in bank_matched:
        adj["银行转账"]["bank"] += brec["_amt"]
for r in third:
    oid = r["商户订单号"].strip()
    pm = "退款" if r["状态"] == "退款" else (erp_by_oid[oid]["支付方式"] if oid in erp_by_oid else "未知")
    adj[pm]["third"] += r["_amt"]
    adj[pm]["fee"] += r["_fee"]


# ═══════════════════════════════════════════
# 5. 输出 Excel
# ═══════════════════════════════════════════

wb = Workbook()
hfont = Font(bold=True, size=11, color="FFFFFF")
hfill = PatternFill("solid", fgColor="4472C4")
halign = Alignment(horizontal="center", vertical="center", wrap_text=True)
calign = Alignment(vertical="center", wrap_text=True)
bd = Border(*(Side(style="thin", color="D9D9D9"),) * 4)
err_bg = PatternFill("solid", fgColor="FCE4EC")
warn_bg = PatternFill("solid", fgColor="FFF2CC")

def hdr(ws, r, cols):
    for i, c in enumerate(cols, 1):
        cell = ws.cell(row=r, column=i, value=c)
        cell.font = hfont; cell.fill = hfill; cell.alignment = halign; cell.border = bd

def bcell(ws, r, c, v, fmt=None):
    cell = ws.cell(row=r, column=c, value=v)
    cell.alignment = calign; cell.border = bd
    if fmt: cell.number_format = fmt

# ── Sheet 1: 对账汇总 ──
ws = wb.active
ws.title = "对账汇总"
ws.cell(row=1, column=1, value="财务多系统对账报告 — 2026年4月").font = Font(bold=True, size=14)
ws.merge_cells("A1:E1")

hdr(ws, 3, ["指标", "数量", "涉及ERP金额", "涉及银行金额", "涉及第三方金额"])
rows = [
    ("总记录数", len(erp), len(bank), len(third), ""),
]
for t, d in sorted(type_summary.items()):
    rows.append((t, d["n"], d["ea"], d["ba"], d["ta"]))
rows.append(("差异总计", len(diffs), "", "", ""))

for ri, (label, n, ea, ba, ta) in enumerate(rows, 4):
    bcell(ws, ri, 1, label)
    bcell(ws, ri, 2, n)
    bcell(ws, ri, 3, ea if isinstance(ea, (int, float)) else ea, "#,##0.00")
    bcell(ws, ri, 4, ba if isinstance(ba, (int, float)) else ba, "#,##0.00")
    bcell(ws, ri, 5, ta if isinstance(ta, (int, float)) else ta, "#,##0.00")
ws.column_dimensions["A"].width = 24
for c in "BCDE": ws.column_dimensions[c].width = 18

# ── Sheet 2: 差异明细 ──
ws2 = wb.create_sheet("差异明细")
hdr(ws2, 1, ["差异类型", "数据来源", "参考号", "ERP金额", "银行金额", "第三方金额", "差异说明"])
for ri, d in enumerate(diffs, 2):
    for ci, k in enumerate(["差异类型", "数据来源", "参考号", "ERP金额", "银行金额", "第三方金额", "差异说明"], 1):
        v = d[k]; bcell(ws2, ri, ci, v, "#,##0.00" if k.endswith("金额") else None)
    if "金额差异" in d["差异类型"]:
        for c in range(1, 8): ws2.cell(row=ri, column=c).fill = err_bg
    elif any(x in d["差异类型"] for x in ["缺失", "缺少"]):
        for c in range(1, 8): ws2.cell(row=ri, column=c).fill = warn_bg
for c, w in zip("ABCDEFG", [18, 14, 22, 14, 14, 14, 65]):
    ws2.column_dimensions[c].width = w

# ── Sheet 3: 调节表 ──
ws3 = wb.create_sheet("调节表")
ws3.cell(row=1, column=1, value="按支付方式调节表 — 2026年4月").font = Font(bold=True, size=14)
ws3.merge_cells("A1:G1")
hdr(ws3, 3, ["支付方式", "ERP应收", "银行实收", "第三方收款", "手续费", "第三方净额", "差异"])
ri = 4
for pm in ["银行转账", "支付宝", "微信支付", "退款"]:
    d = adj[pm]
    net3 = round(d["third"] - d["fee"], 2)
    received = round(d["bank"] + net3, 2)
    diff = round(d["erp"] - received, 2)
    for ci, v in enumerate([pm, round(d["erp"],2), round(d["bank"],2), round(d["third"],2), round(d["fee"],2), net3, diff], 1):
        bcell(ws3, ri, ci, v, "#,##0.00" if ci > 1 else None)
    ri += 1
# 合计
for ci in range(2, 8):
    vals = [ws3.cell(row=r, column=ci).value for r in range(4, ri)]
    total = round(sum(v for v in vals if isinstance(v, (int, float))), 2)
    bcell(ws3, ri, ci, total, "#,##0.00")
bcell(ws3, ri, 1, "合计"); ws3.cell(row=ri, column=1).font = Font(bold=True)
for c in "ABCDEFG": ws3.column_dimensions[c].width = 16
ri += 2
for note in ["差异 = ERP应收 - (银行实收 + 第三方收款 - 手续费)", "正数: ERP应收多于到账 | 负数: 到账多于ERP应收", "第三方已扣除手续费"]:
    bcell(ws3, ri, 1, note); ri += 1

# 保存
path = os.path.join(OUTPUT_DIR, "财务多系统对账报告_2026年4月.xlsx")
wb.save(path)
print(f"\n生成: {path}")
print(f"差异总数: {len(diffs)}")
for t, d in sorted(type_summary.items()):
    print(f"  {t}: {d['n']} 条")
