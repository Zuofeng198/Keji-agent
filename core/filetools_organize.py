"""文件整理 & 数据处理工具 —— 分类归档/批量重命名/去重/ETL/清洗/转换"""

import os
import re
import json
import csv
import io
import hashlib
import datetime
import shutil
from typing import Optional

from core.tools import register_tool


def _format_size(size: int) -> str:
    for unit in ["B", "KB", "MB", "GB"]:
        if size < 1024:
            return f"{size:.1f}{unit}"
        size /= 1024
    return f"{size:.1f}TB"


def _get_file_category(ext: str) -> str:
    """根据扩展名返回文件分类"""
    ext = ext.lower()
    if ext in (".jpg", ".jpeg", ".png", ".gif", ".bmp", ".tiff", ".tif", ".webp", ".svg", ".ico"):
        return "图片"
    if ext in (".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx", ".pdf", ".txt", ".md", ".csv"):
        return "文档"
    if ext in (".mp4", ".avi", ".mov", ".wmv", ".flv", ".mkv", ".webm", ".mpg", ".mpeg"):
        return "视频"
    if ext in (".mp3", ".wav", ".flac", ".aac", ".ogg", ".wma", ".m4a"):
        return "音频"
    if ext in (".zip", ".rar", ".7z", ".tar", ".gz", ".bz2", ".xz"):
        return "压缩包"
    if ext in (".py", ".js", ".ts", ".java", ".cpp", ".c", ".h", ".go", ".rs", ".php", ".rb", ".swift"):
        return "代码"
    if ext in (".exe", ".msi", ".dll", ".deb", ".rpm", ".appimage"):
        return "可执行文件"
    return "其他"


# ═══════════════════════════════════════════════════════════════
# 工具：文件分类整理
# ═══════════════════════════════════════════════════════════════

@register_tool(
    name="organize_files",
    description="按类型/日期/大小自动归类文件到分类文件夹中，支持按扩展名分组或自定义规则",
    parameters={
        "source_dir": {"type": "string", "description": "要整理的文件夹路径"},
        "mode": {
            "type": "string",
            "description": "分类模式：type（按文件类型）、date（按年月）、size（按大小）、flat（只列出分类预览不移动）",
        },
        "preview": {
            "type": "boolean",
            "description": "预览模式（只显示分类结果不移动文件），默认 false",
        },
        "recursive": {
            "type": "boolean",
            "description": "是否递归处理子文件夹，默认 false",
        },
    },
    category="filesystem",
    timeout=120,
)
def organize_files(source_dir: str = "", mode: str = "type", preview: bool = False, recursive: bool = False) -> str:
    from core.path_policy import check_path, default_browse_path
    if not source_dir:
        source_dir = default_browse_path()
    source_dir, err = check_path(source_dir, must_exist=True, must_be_dir=True)
    if err:
        return err

    if not os.path.isdir(source_dir):
        return f"错误：文件夹不存在「{source_dir}」"

    # 收集文件
    files = []
    if recursive:
        for root, dirs, fnames in os.walk(source_dir):
            for fn in fnames:
                files.append(os.path.join(root, fn))
    else:
        for fn in os.listdir(source_dir):
            fp = os.path.join(source_dir, fn)
            if os.path.isfile(fp):
                files.append(fp)

    if not files:
        return "文件夹中没有文件"

    files.sort()

    # 分类
    categorized = {}
    for fp in files:
        try:
            stat = os.stat(fp)
            ext = os.path.splitext(fp)[1].lower()
            fname = os.path.basename(fp)
            mtime = datetime.datetime.fromtimestamp(stat.st_mtime)
            size = stat.st_size
        except Exception:
            continue

        if mode == "type":
            cat = _get_file_category(ext)
        elif mode == "date":
            cat = mtime.strftime("%Y年%m月")
        elif mode == "size":
            if size < 100 * 1024:
                cat = "小于100KB"
            elif size < 1024 * 1024:
                cat = "100KB~1MB"
            elif size < 100 * 1024 * 1024:
                cat = "1MB~100MB"
            else:
                cat = "大于100MB"
        else:  # flat
            cat = _get_file_category(ext)

        categorized.setdefault(cat, []).append(fp)

    # 输出结果
    lines = [f"📂 整理文件夹: {source_dir}", f"模式: {mode} | 共 {len(files)} 个文件 | 分为 {len(categorized)} 类\n"]

    for cat in sorted(categorized):
        items = categorized[cat]
        total_size = sum(os.path.getsize(f) for f in items if os.path.isfile(f))
        lines.append(f"── {cat} ({len(items)} 个, {_format_size(total_size)}) ──")
        for fp in items:
            fname = os.path.basename(fp)
            fsize = _format_size(os.path.getsize(fp))
            lines.append(f"  📄 {fname} ({fsize})")
        lines.append("")

    if not preview and mode != "flat":
        # 实际移动文件
        moved = 0
        for cat in sorted(categorized):
            target_dir = os.path.join(source_dir, cat)
            os.makedirs(target_dir, exist_ok=True)

            for fp in categorized[cat]:
                # 如果文件已经在目标目录，跳过
                if os.path.dirname(fp) == target_dir:
                    continue
                try:
                    dest = os.path.join(target_dir, os.path.basename(fp))
                    # 重名处理
                    if os.path.exists(dest):
                        base, ext = os.path.splitext(os.path.basename(fp))
                        counter = 1
                        while os.path.exists(os.path.join(target_dir, f"{base}_{counter}{ext}")):
                            counter += 1
                        dest = os.path.join(target_dir, f"{base}_{counter}{ext}")
                    shutil.move(fp, dest)
                    moved += 1
                except Exception as e:
                    lines.append(f"  ❌ 移动失败: {os.path.basename(fp)} - {str(e)[:60]}")

        lines.append(f"\n✅ 已移动 {moved} 个文件到分类文件夹")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
# 工具：批量重命名
# ═══════════════════════════════════════════════════════════════

@register_tool(
    name="rename_files",
    description="批量重命名文件，支持添加前缀/后缀、替换文字、正则替换、序号编号",
    parameters={
        "directory": {"type": "string", "description": "要重命名的文件所在文件夹路径"},
        "pattern": {
            "type": "string",
            "description": "操作模式：prefix（加前缀）、suffix（加后缀）、replace（替换文字）、regex（正则替换）、number（序号编号）",
        },
        "value": {
            "type": "string",
            "description": "模式参数：prefix/suffix 时是要加的文字，replace 时是 旧文字|新文字，regex 时是 正则|替换, number 时是 前缀文字",
        },
        "ext_filter": {
            "type": "string",
            "description": "扩展名过滤（只处理指定类型的文件），如 .txt,.md 或留空处理所有",
        },
        "preview": {
            "type": "boolean",
            "description": "预览模式（只显示重命名结果不实际执行），默认 true",
        },
        "start_from": {
            "type": "integer",
            "description": "编号起始值（仅 number 模式有效），默认 1",
        },
    },
    category="filesystem",
    timeout=60,
)
def rename_files(directory: str = "", pattern: str = "prefix", value: str = "",
                 ext_filter: str = "", preview: bool = True, start_from: int = 1) -> str:
    from core.path_policy import check_path, default_browse_path
    if not directory:
        directory = default_browse_path()
    directory, err = check_path(directory, must_exist=True, must_be_dir=True)
    if err:
        return err

    if not os.path.isdir(directory):
        return f"错误：文件夹不存在「{directory}」"

    # 收集文件
    all_files = []
    for fn in sorted(os.listdir(directory)):
        fp = os.path.join(directory, fn)
        if not os.path.isfile(fp):
            continue
        # 扩展名过滤
        if ext_filter:
            allowed = [e.strip().lower() for e in ext_filter.split(",")]
            if os.path.splitext(fn)[1].lower() not in allowed:
                continue
        all_files.append(fn)

    if not all_files:
        return "没有找到符合条件的文件"

    # 生成新文件名
    renamed = []
    for i, fn in enumerate(all_files):
        base, ext = os.path.splitext(fn)
        new_name = fn  # 默认不变

        try:
            if pattern == "prefix":
                new_name = value + fn
            elif pattern == "suffix":
                new_name = base + value + ext
            elif pattern == "replace":
                if "|" in value:
                    old_text, new_text = value.split("|", 1)
                    new_name = fn.replace(old_text, new_text)
                else:
                    return "错误：replace 模式格式为 旧文字|新文字"
            elif pattern == "regex":
                if "|" in value:
                    reg, repl = value.split("|", 1)
                    new_name = re.sub(reg, repl, fn)
                else:
                    return "错误：regex 模式格式为 正则表达式|替换文字"
            elif pattern == "number":
                num = start_from + i
                digits = len(str(start_from + len(all_files) - 1))
                prefix_text = value
                new_name = f"{prefix_text}{num:0{digits}d}{ext}"
            else:
                return f"错误：未知模式「{pattern}」，支持 prefix/suffix/replace/regex/number"
        except re.error as e:
            return f"正则表达式错误: {e}"
        except Exception as e:
            renamed.append((fn, new_name, f"❌ {str(e)[:60]}"))
            continue

        if fn != new_name:
            renamed.append((fn, new_name, ""))

    if not renamed:
        return "所有文件名已符合规则，无需修改"

    lines = [f"📝 批量重命名 - {pattern}模式", f"文件夹: {directory}", f"匹配文件: {len(all_files)} 个 | 需改名: {len(renamed)} 个\n"]

    for old_name, new_name, err in renamed:
        status = err if err else "✅"
        lines.append(f"  {old_name}  →  {new_name}  {status}")

    if not preview and not any(err for _, _, err in renamed):
        actual_move = 0
        for old_name, new_name, err in renamed:
            if err:
                continue
            try:
                old_path = os.path.join(directory, old_name)
                new_path = os.path.join(directory, new_name)
                if os.path.exists(new_path):
                    base, ext = os.path.splitext(new_name)
                    counter = 1
                    while os.path.exists(os.path.join(directory, f"{base}_{counter}{ext}")):
                        counter += 1
                    new_path = os.path.join(directory, f"{base}_{counter}{ext}")
                os.rename(old_path, new_path)
                actual_move += 1
            except Exception as e:
                lines.append(f"  ❌ 重命名失败: {old_name} - {str(e)[:60]}")
        lines.append(f"\n✅ 已重命名 {actual_move} 个文件")

    if preview:
        lines.append(f"\n💡 预览模式，未实际执行。确认无误后设置 preview=false 再运行一次")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
# 工具：文件去重
# ═══════════════════════════════════════════════════════════════

@register_tool(
    name="deduplicate_files",
    description="查找并删除重复文件（基于 MD5 哈希），支持模糊名称匹配",
    parameters={
        "directory": {"type": "string", "description": "要扫描的文件夹路径"},
        "action": {
            "type": "string",
            "description": "操作：scan（只扫描列出重复）、delete（删除重复文件，保留一个）、move_to（移动到指定目录）",
        },
        "move_dir": {
            "type": "string",
            "description": "重复文件移动目标目录（仅 move_to 模式有效）",
        },
        "fuzzy_name": {
            "type": "boolean",
            "description": "是否启用文件名模糊匹配（如 报告(1).docx 和 报告.docx），默认 true",
        },
    },
    category="filesystem",
    timeout=120,
)
def deduplicate_files(directory: str = "", action: str = "scan",
                      move_dir: str = "", fuzzy_name: bool = True) -> str:
    from core.path_policy import check_path, default_browse_path
    if not directory:
        directory = default_browse_path()
    directory, err = check_path(directory, must_exist=True, must_be_dir=True)
    if err:
        return err

    if not os.path.isdir(directory):
        return f"错误：文件夹不存在「{directory}」"

    # 递归扫描所有文件
    all_files = []
    for root, dirs, fnames in os.walk(directory):
        for fn in fnames:
            fp = os.path.join(root, fn)
            try:
                stat = os.stat(fp)
                all_files.append({
                    "path": fp,
                    "name": fn,
                    "size": stat.st_size,
                    "mtime": stat.st_mtime,
                })
            except Exception:
                pass

    if not all_files:
        return "文件夹中没有文件"

    # 按大小分组（先快速筛选）
    size_groups = {}
    for f in all_files:
        size_groups.setdefault(f["size"], []).append(f)

    # 对同大小文件计算 MD5
    md5_groups = {}
    fuzzy_groups = {}
    for size, files in size_groups.items():
        if len(files) < 2:
            continue
        for f in files:
            try:
                md5 = hashlib.md5()
                with open(f["path"], "rb") as fp:
                    for chunk in iter(lambda: fp.read(8192), b""):
                        md5.update(chunk)
                f["md5"] = md5.hexdigest()
                md5_groups.setdefault(md5.hexdigest(), []).append(f)
            except Exception:
                pass

    # 精确 MD5 重复
    exact_dupes = {k: v for k, v in md5_groups.items() if len(v) > 1}

    # 模糊名称匹配
    fuzzy_dupes = []
    if fuzzy_name:
        for f in all_files:
            base = os.path.splitext(f["name"])[0]
            # 匹配 "(1)" 或 "_1" 或 " - 副本" 这类重复标记
            fuzzy_base = re.sub(r"[\s\-_]*\(?\d+\)?$|[\s\-_]*副本[\s\-_]*$", "", base).strip().lower()
            if fuzzy_base:
                f["fuzzy_key"] = fuzzy_base

        fuzzy_map = {}
        for f in all_files:
            if "fuzzy_key" not in f:
                continue
            fuzzy_map.setdefault(f["fuzzy_key"], []).append(f)
        fuzzy_dupes = {k: v for k, v in fuzzy_map.items() if len(v) > 1 and k != ""}

    # 输出结果
    lines = [f"🔍 文件去重扫描", f"文件夹: {directory}", f"总文件: {len(all_files)} 个\n"]

    if exact_dupes:
        lines.append(f"📊 MD5 完全重复组: {len(exact_dupes)} 组")
        total_saved = 0
        for md5, files in sorted(exact_dupes.items()):
            size = _format_size(files[0]["size"])
            lines.append(f"\n  ── MD5: {md5[:12]}... ({size}, {len(files)} 个文件) ──")
            for i, f in enumerate(files):
                keep = " ✅ 保留" if i == 0 else " ❌ 可删除"
                mtime = datetime.datetime.fromtimestamp(f["mtime"]).strftime("%Y-%m-%d %H:%M")
                lines.append(f"    {f['name']}{keep}")
                if i > 0:
                    total_saved += f["size"]
            lines.append(f"  可释放空间: {_format_size(total_saved)}")
    else:
        lines.append("✅ 未发现 MD5 完全重复的文件")

    if fuzzy_dupes and fuzzy_name:
        lines.append(f"\n📊 模糊名称相似组: {len(fuzzy_dupes)} 组")
        for key, files in sorted(fuzzy_dupes.items()):
            if len(files) < 2:
                continue
            lines.append(f"\n  ── 相似名: {key} ({len(files)} 个) ──")
            for f in files:
                lines.append(f"    {f['name']} ({_format_size(f['size'])})")

    # 执行操作
    if action == "delete" and exact_dupes:
        deleted = 0
        saved = 0
        for md5, files in exact_dupes.items():
            for i, f in enumerate(files):
                if i == 0:
                    continue  # 保留第一个
                try:
                    os.remove(f["path"])
                    deleted += 1
                    saved += f["size"]
                except Exception as e:
                    lines.append(f"\n  ❌ 删除失败: {f['name']} - {str(e)[:60]}")
        lines.append(f"\n✅ 已删除 {deleted} 个重复文件，释放 {_format_size(saved)} 空间")

    elif action == "move_to" and exact_dupes and move_dir:
        move_dir, err = check_path(move_dir)
        if err:
            return err
        os.makedirs(move_dir, exist_ok=True)
        moved = 0
        for md5, files in exact_dupes.items():
            for i, f in enumerate(files):
                if i == 0:
                    continue
                try:
                    dest = os.path.join(move_dir, f["name"])
                    if os.path.exists(dest):
                        base, ext = os.path.splitext(f["name"])
                        c = 1
                        while os.path.exists(os.path.join(move_dir, f"{base}_{c}{ext}")):
                            c += 1
                        dest = os.path.join(move_dir, f"{base}_{c}{ext}")
                    shutil.move(f["path"], dest)
                    moved += 1
                except Exception as e:
                    lines.append(f"\n  ❌ 移动失败: {f['name']} - {str(e)[:60]}")
        lines.append(f"\n✅ 已移动 {moved} 个重复文件到 {move_dir}")

    return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
# 工具：ETL 数据处理管道
# ═══════════════════════════════════════════════════════════════

@register_tool(
    name="etl_pipeline",
    description="ETL 数据处理管道：从文件提取数据→转换处理→输出结果，支持链式操作",
    parameters={
        "source": {"type": "string", "description": "数据源：文件路径或直接输入的数据内容"},
        "steps": {
            "type": "string",
            "description": "处理步骤（JSON 数组），每步为一个对象：{\"op\":\"操作\",\"params\":{}}。支持过滤 filter、排序 sort、分组 group、聚合 aggregate、列操作 columns、去重 distinct、空值处理 fillna、转换类型 cast",
        },
        "output_format": {
            "type": "string",
            "description": "输出格式：table（表格文字）、json、csv、xlsx，默认 table",
        },
        "output_path": {
            "type": "string",
            "description": "输出文件路径（仅 xlsx/csv 格式需要）",
        },
    },
    category="data",
    timeout=60,
)
def etl_pipeline(source: str = "", steps: str = "",
                 output_format: str = "table", output_path: str = "") -> str:
    """执行 ETL 数据处理管道"""
    import ast
    import statistics as stats_mod

    # ---- 1. Extract：读取数据 ----
    rows = []
    if os.path.isfile(source):
        ext = os.path.splitext(source)[1].lower()
        try:
            if ext == ".csv":
                with open(source, "r", encoding="utf-8") as f:
                    rows = list(csv.reader(f))
            elif ext == ".xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(source, read_only=True, data_only=True)
                ws = wb.active
                rows = [[str(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
            elif ext == ".json":
                with open(source, "r", encoding="utf-8") as f:
                    data = json.load(f)
                    if isinstance(data, list) and data:
                        if isinstance(data[0], dict):
                            headers = list(data[0].keys())
                            rows = [headers] + [[str(r.get(h, "")) for h in headers] for r in data]
                        else:
                            rows = [["value"]] + [[str(v)] for v in data]
            else:
                with open(source, "r", encoding="utf-8") as f:
                    text = f.read()
                    rows = [line.split(",") for line in text.strip().split("\n") if line.strip()]
        except Exception as e:
            return f"读取数据源失败: {e}"
    else:
        for line in source.strip().split("\n"):
            line = line.strip()
            if line:
                rows.append([c.strip() for c in line.split(",")])

    if not rows or len(rows) < 1:
        return "数据为空"

    headers = rows[0] if rows else []
    data_rows = rows[1:] if len(rows) > 1 else []

    # ---- 2. Transform：执行处理步骤 ----
    if steps:
        try:
            step_list = json.loads(steps) if isinstance(steps, str) else steps
        except json.JSONDecodeError:
            return "错误：steps 参数必须是有效的 JSON 数组"

        for step in step_list:
            op = step.get("op", "")
            params = step.get("params", {})

            try:
                if op == "filter":
                    col = params.get("column", 0)
                    op_fn = params.get("op", "eq")
                    val = params.get("value", "")
                    col_idx = _resolve_col(col, headers)

                    new_rows = []
                    for r in data_rows:
                        cell = r[col_idx] if col_idx < len(r) else ""
                        try:
                            cell_num = float(cell.replace(",", "").replace("¥", "").replace("$", ""))
                            val_num = float(val)
                            if op_fn == "gt" and cell_num > val_num: new_rows.append(r)
                            elif op_fn == "lt" and cell_num < val_num: new_rows.append(r)
                            elif op_fn == "ge" and cell_num >= val_num: new_rows.append(r)
                            elif op_fn == "le" and cell_num <= val_num: new_rows.append(r)
                            elif op_fn == "eq" and cell_num == val_num: new_rows.append(r)
                            elif op_fn == "ne" and cell_num != val_num: new_rows.append(r)
                            else: pass
                        except (ValueError, IndexError):
                            if op_fn == "eq" and cell == val: new_rows.append(r)
                            elif op_fn == "ne" and cell != val: new_rows.append(r)
                            elif op_fn == "contains" and val in cell: new_rows.append(r)
                            elif op_fn == "not_contains" and val not in cell: new_rows.append(r)
                            else: pass
                    data_rows = new_rows

                elif op == "sort":
                    col = params.get("column", 0)
                    desc = params.get("desc", False)
                    col_idx = _resolve_col(col, headers)
                    data_rows.sort(key=lambda r: _sort_key(r, col_idx), reverse=desc)

                elif op == "distinct":
                    seen = set()
                    new_rows = []
                    for r in data_rows:
                        key = tuple(r)
                        if key not in seen:
                            seen.add(key)
                            new_rows.append(r)
                    data_rows = new_rows

                elif op == "columns":
                    cols = params.get("columns", [])
                    indices = [_resolve_col(c, headers) for c in cols]
                    if indices:
                        headers = [headers[i] for i in indices]
                        data_rows = [[r[i] for i in indices] for r in data_rows]

                elif op == "fillna":
                    col = params.get("column", 0)
                    fill_val = params.get("value", "")
                    col_idx = _resolve_col(col, headers)
                    for r in data_rows:
                        if col_idx < len(r) and (r[col_idx] == "" or r[col_idx] is None):
                            r[col_idx] = str(fill_val)

                elif op == "aggregate":
                    col = params.get("column", 0)
                    agg_fn = params.get("function", "count")
                    col_idx = _resolve_col(col, headers)
                    values = []
                    for r in data_rows:
                        try:
                            v = float(r[col_idx].replace(",", "").replace("¥", "").replace("$", ""))
                            values.append(v)
                        except (ValueError, IndexError):
                            pass

                    aggs = {"sum": sum(values), "avg": stats_mod.mean(values) if values else 0,
                            "max": max(values) if values else 0, "min": min(values) if values else 0,
                            "count": len(values), "std": stats_mod.stdev(values) if len(values) > 1 else 0}
                    agg_name = params.get("name", f"{agg_fn}_{col}")

                    # 添加聚合行
                    result_row = [""] * len(headers)
                    if isinstance(col, str) and col in headers:
                        result_col = _resolve_col(col, headers)
                        result_row[result_col] = aggs.get(agg_fn, "")
                    elif isinstance(col, int) and col < len(headers):
                        result_row[col] = aggs.get(agg_fn, "")
                    else:
                        result_row[0] = str(aggs.get(agg_fn, ""))
                    data_rows.append(result_row)

                elif op == "group":
                    by_col = params.get("by", 0)
                    agg_col = params.get("agg_column", 1)
                    agg_fn = params.get("function", "sum")
                    by_idx = _resolve_col(by_col, headers)
                    agg_idx = _resolve_col(agg_col, headers)

                    groups = {}
                    for r in data_rows:
                        key = r[by_idx] if by_idx < len(r) else ""
                        val = 0
                        try:
                            val = float(r[agg_idx].replace(",", "").replace("¥", "").replace("$", ""))
                        except (ValueError, IndexError):
                            val = 0
                        groups.setdefault(key, []).append(val)

                    result_rows = []
                    for key, vals in groups.items():
                        if agg_fn == "sum": res = sum(vals)
                        elif agg_fn == "avg": res = stats_mod.mean(vals)
                        elif agg_fn == "max": res = max(vals)
                        elif agg_fn == "min": res = min(vals)
                        elif agg_fn == "count": res = len(vals)
                        else: res = sum(vals)
                        result_rows.append([key, str(round(res, 2))])

                    # 更新表头和数据
                    if by_idx < len(headers) and agg_idx < len(headers):
                        headers = [headers[by_idx], f"{agg_fn}({headers[agg_idx]})"]
                    else:
                        headers = ["分组", f"{agg_fn}"]
                    data_rows = result_rows

            except Exception as e:
                return f"步骤 [{op}] 执行失败: {str(e)[:200]}"

    # ---- 3. Load：输出结果 ----
    if output_format == "json":
        result_list = []
        for r in data_rows:
            row_dict = {}
            for i, h in enumerate(headers):
                row_dict[h] = r[i] if i < len(r) else ""
            result_list.append(row_dict)
        return json.dumps(result_list, ensure_ascii=False, indent=2)

    elif output_format in ("csv", "xlsx"):
        if output_format == "csv":
            if not output_path:
                output_path = os.path.expanduser("~\\Desktop\\etl_output.csv")
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                writer = csv.writer(f)
                writer.writerow(headers)
                writer.writerows(data_rows)
            return f"✅ CSV 已保存: {output_path} ({len(data_rows)} 行)"

        else:  # xlsx
            if not output_path:
                output_path = os.path.expanduser("~\\Desktop\\etl_output.xlsx")
            try:
                from openpyxl import Workbook
                wb = Workbook()
                ws = wb.active
                ws.title = "ETL结果"
                ws.append(headers)
                for r in data_rows:
                    ws.append(r)
                wb.save(output_path)
                return f"✅ Excel 已保存: {output_path} ({len(data_rows)} 行)"
            except ImportError:
                return "错误：需要 openpyxl 库"

    else:  # table
        lines = []
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for r in data_rows:
            line = []
            for i, h in enumerate(headers):
                line.append(r[i] if i < len(r) else "")
            lines.append("| " + " | ".join(line) + " |")
        lines.insert(0, f"📊 ETL 处理结果 ({len(data_rows)} 行)")
        return "\n".join(lines)


def _resolve_col(col, headers):
    """将列名或列号转为列索引"""
    if isinstance(col, str) and col.isdigit():
        return int(col)
    if isinstance(col, str):
        col_lower = col.lower()
        for i, h in enumerate(headers):
            if col_lower in h.lower():
                return i
        return 0
    return int(col) if col < len(headers) else 0


def _sort_key(row, col_idx):
    """排序键函数（数值优先，文本回退）"""
    if col_idx < len(row):
        try:
            return (0, float(row[col_idx].replace(",", "").replace("¥", "").replace("$", "")))
        except (ValueError, IndexError):
            return (1, row[col_idx])
    return (1, "")


# ═══════════════════════════════════════════════════════════════
# 工具：数据清洗
# ═══════════════════════════════════════════════════════════════

@register_tool(
    name="clean_data",
    description="清洗数据：处理空值、统一格式、去除异常值、去除空格、去除重复行",
    parameters={
        "source": {"type": "string", "description": "数据源：文件路径（CSV/Excel）或直接输入的CSV文本"},
        "operations": {
            "type": "string",
            "description": "清洗操作，逗号分隔：trim（去空格）、fillna（填充空值）、dropna（删除空行）、dedup（去重行）、strip_punctuation（去标点）、lowercase（转小写）、uppercase（转大写）、normalize_whitespace（合并空格）",
        },
        "fill_value": {
            "type": "string",
            "description": "填充空值的默认值，仅 fillna 操作有效，默认 'N/A'",
        },
        "columns": {
            "type": "string",
            "description": "指定操作的列，逗号分隔列名或列号（默认全部列）",
        },
        "output_format": {
            "type": "string",
            "description": "输出格式：table、json、csv，默认 table",
        },
    },
    category="data",
    timeout=30,
)
def clean_data(source: str = "", operations: str = "trim",
               fill_value: str = "N/A", columns: str = "",
               output_format: str = "table") -> str:
    rows = []
    if os.path.isfile(source):
        ext = os.path.splitext(source)[1].lower()
        try:
            if ext == ".csv":
                with open(source, "r", encoding="utf-8") as f:
                    rows = list(csv.reader(f))
            elif ext == ".xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(source, read_only=True, data_only=True)
                ws = wb.active
                rows = [[str(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
            else:
                return f"不支持的文件类型: {ext}"
        except Exception as e:
            return f"读取文件失败: {e}"
    else:
        for line in source.strip().split("\n"):
            line = line.strip()
            if line:
                rows.append([c.strip() for c in line.split(",")])

    if not rows:
        return "无数据"

    headers = rows[0]
    data_rows = rows[1:]
    if not data_rows:
        return "仅有表头，无数据行"

    # 确定操作的列
    if columns:
        col_indices = []
        for c in columns.split(","):
            c = c.strip()
            if c.isdigit():
                col_indices.append(int(c))
            else:
                for i, h in enumerate(headers):
                    if c.lower() in h.lower():
                        col_indices.append(i)
        col_indices = list(set(col_indices))
    else:
        col_indices = list(range(len(headers)))

    ops = [o.strip().lower() for o in operations.split(",") if o.strip()]

    summary = []
    original_count = len(data_rows)

    for op in ops:
        if op == "trim":
            for r in data_rows:
                for ci in col_indices:
                    if ci < len(r):
                        r[ci] = r[ci].strip()
            summary.append(f"trim: 去除 {len(col_indices)} 列空格")

        elif op == "fillna":
            filled = 0
            for r in data_rows:
                for ci in col_indices:
                    if ci < len(r) and (r[ci] == "" or r[ci] is None):
                        r[ci] = fill_value
                        filled += 1
            summary.append(f"fillna: 填充 {filled} 个空值")

        elif op == "dropna":
            new_rows = []
            dropped = 0
            for r in data_rows:
                is_empty = all(ci < len(r) and (r[ci] == "" or r[ci] is None) for ci in col_indices)
                if is_empty:
                    dropped += 1
                else:
                    new_rows.append(r)
            data_rows = new_rows
            summary.append(f"dropna: 删除 {dropped} 个空行")

        elif op == "dedup":
            seen = set()
            new_rows = []
            duped = 0
            for r in data_rows:
                key = tuple(r[ci] if ci < len(r) else "" for ci in col_indices)
                if key in seen:
                    duped += 1
                else:
                    seen.add(key)
                    new_rows.append(r)
            data_rows = new_rows
            summary.append(f"dedup: 去重 {duped} 行")

        elif op == "lowercase":
            for r in data_rows:
                for ci in col_indices:
                    if ci < len(r):
                        r[ci] = r[ci].lower()
            summary.append(f"lowercase: {len(col_indices)} 列转小写")

        elif op == "uppercase":
            for r in data_rows:
                for ci in col_indices:
                    if ci < len(r):
                        r[ci] = r[ci].upper()
            summary.append(f"uppercase: {len(col_indices)} 列转大写")

        elif op == "strip_punctuation":
            import string
            punct = string.punctuation + "，。、；：？！＂＃＄％＆＇（）＊＋，－．／：；＜＝＞？＠［＼］＾＿｀｛｜｝～｟｠｢｣"
            for r in data_rows:
                for ci in col_indices:
                    if ci < len(r):
                        r[ci] = r[ci].strip(punct)
            summary.append(f"strip_punctuation: 去除标点")

        elif op == "normalize_whitespace":
            for r in data_rows:
                for ci in col_indices:
                    if ci < len(r):
                        r[ci] = re.sub(r"\s+", " ", r[ci]).strip()
            summary.append(f"normalize_whitespace: 合并多余空格")

    # 输出
    summary_str = f"清洗前 {original_count} 行 → 清洗后 {len(data_rows)} 行\n" + "\n".join(f"  ✅ {s}" for s in summary)

    if output_format == "json":
        result = []
        for r in data_rows:
            d = {}
            for i, h in enumerate(headers):
                d[h] = r[i] if i < len(r) else ""
            result.append(d)
        return f"{summary_str}\n\n{json.dumps(result, ensure_ascii=False, indent=2)}"

    elif output_format == "csv":
        output = io.StringIO()
        w = csv.writer(output)
        w.writerow(headers)
        w.writerows(data_rows)
        return f"{summary_str}\n\n{output.getvalue()}"

    else:
        lines = [f"🧹 数据清洗结果\n{summary_str}\n"]
        lines.append("| " + " | ".join(headers) + " |")
        lines.append("| " + " | ".join(["---"] * len(headers)) + " |")
        for r in data_rows[:50]:
            lines.append("| " + " | ".join(r[i] if i < len(r) else "" for i in range(len(headers))) + " |")
        if len(data_rows) > 50:
            lines.append(f"\n... 仅显示前 50 行，共 {len(data_rows)} 行")
        return "\n".join(lines)


# ═══════════════════════════════════════════════════════════════
# 工具：格式批量转换
# ═══════════════════════════════════════════════════════════════

@register_tool(
    name="convert_data",
    description="批量转换数据格式：CSV ↔ Excel ↔ JSON ↔ HTML 表格互转",
    parameters={
        "source": {"type": "string", "description": "源文件路径（支持 .csv/.xlsx/.json）或直接输入数据"},
        "target_format": {
            "type": "string",
            "description": "目标格式：csv、xlsx、json、html",
        },
        "output_path": {
            "type": "string",
            "description": "输出文件路径（默认保存到桌面）",
        },
    },
    category="data",
    timeout=30,
)
def convert_data(source: str = "", target_format: str = "csv", output_path: str = "") -> str:
    rows = []
    source_is_file = os.path.isfile(source)

    if source_is_file:
        ext = os.path.splitext(source)[1].lower()
        try:
            if ext == ".csv":
                with open(source, "r", encoding="utf-8") as f:
                    rows = list(csv.reader(f))
            elif ext == ".xlsx":
                from openpyxl import load_workbook
                wb = load_workbook(source, read_only=True, data_only=True)
                ws = wb.active
                rows = [[str(c) if c is not None else "" for c in row] for row in ws.iter_rows(values_only=True)]
            elif ext == ".json":
                with open(source, "r", encoding="utf-8") as f:
                    data = json.load(f)
                if isinstance(data, list) and data and isinstance(data[0], dict):
                    headers = list(data[0].keys())
                    rows = [headers] + [[str(r.get(h, "")) for h in headers] for r in data]
                elif isinstance(data, dict):
                    headers = list(data.keys())
                    rows = [headers] + [[str(v) for v in data.values()]]
                else:
                    return "不支持的 JSON 格式"
            else:
                return f"不支持的文件类型: {ext}"
        except Exception as e:
            return f"读取文件失败: {e}"
    else:
        for line in source.strip().split("\n"):
            line = line.strip()
            if line:
                rows.append([c.strip() for c in line.split(",")])

    if not rows:
        return "无数据"

    headers = rows[0]
    data_rows = rows[1:]

    if not output_path and source_is_file:
        base = os.path.splitext(source)[0]
        output_path = f"{base}.{target_format}"
    elif not output_path:
        output_path = os.path.expanduser(f"~\\Desktop\\converted.{target_format}")

    try:
        if target_format == "csv":
            output_path = os.path.splitext(output_path)[0] + ".csv"
            with open(output_path, "w", encoding="utf-8-sig", newline="") as f:
                w = csv.writer(f)
                w.writerow(headers)
                w.writerows(data_rows)
            return f"✅ CSV 已保存: {output_path} ({len(data_rows)} 行)"

        elif target_format == "xlsx":
            output_path = os.path.splitext(output_path)[0] + ".xlsx"
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            wb = Workbook()
            ws = wb.active
            ws.title = "数据"
            # 写表头
            header_font = Font(bold=True, color="FFFFFF", size=11)
            header_fill = PatternFill(start_color="4F6EF7", end_color="4F6EF7", fill_type="solid")
            for j, h in enumerate(headers, 1):
                cell = ws.cell(row=1, column=j, value=h)
                cell.font = header_font
                cell.fill = header_fill
            # 写数据
            for i, r in enumerate(data_rows, 2):
                for j, v in enumerate(r, 1):
                    ws.cell(row=i, column=j, value=v)
            wb.save(output_path)
            return f"✅ Excel 已保存: {output_path} ({len(data_rows)} 行)"

        elif target_format == "json":
            output_path = os.path.splitext(output_path)[0] + ".json"
            result = []
            for r in data_rows:
                d = {}
                for i, h in enumerate(headers):
                    d[h] = r[i] if i < len(r) else ""
                result.append(d)
            with open(output_path, "w", encoding="utf-8") as f:
                json.dump(result, f, ensure_ascii=False, indent=2)
            return f"✅ JSON 已保存: {output_path} ({len(data_rows)} 条记录)"

        elif target_format == "html":
            output_path = os.path.splitext(output_path)[0] + ".html"
            html = ['<!DOCTYPE html><html><meta charset="utf-8"><body>']
            html.append('<table border="1" cellpadding="6" style="border-collapse:collapse;width:100%;font-family:sans-serif">')
            html.append("<thead><tr>" + "".join(f"<th style='background:#4F6EF7;color:#fff'>{h}</th>" for h in headers) + "</tr></thead>")
            html.append("<tbody>")
            for i, r in enumerate(data_rows):
                bg = " style='background:#f5f5f5'" if i % 2 == 0 else ""
                html.append(f"<tr{bg}>" + "".join(f"<td>{r[j] if j < len(r) else ''}</td>" for j in range(len(headers))) + "</tr>")
            html.append("</tbody></table></body></html>")
            with open(output_path, "w", encoding="utf-8") as f:
                f.write("\n".join(html))
            return f"✅ HTML 已保存: {output_path} ({len(data_rows)} 行)"

        else:
            return f"不支持的目标格式: {target_format}"

    except ImportError as e:
        return f"缺少库: {e}"
    except Exception as e:
        return f"转换失败: {str(e)[:200]}"
